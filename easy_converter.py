#!/usr/bin/python
# -*- coding: UTF-8 -*-
import os
import re
from abc import ABC
from abc import abstractmethod
from enum import Enum
from openpyxl import load_workbook
from argparse import ArgumentParser
import inflect


def upper_camel_case(var_name):
    if var_name == '':
        return var_name
    return var_name[0].upper() + var_name[1:]


def split_last_word(var_name):
    splits = re_camel_split.findall(var_name)
    last_word = splits[-1]
    rest_part = ""
    if len(splits) > 1:
        rest_part = "".join(splits[0:-1])
    capitalized = last_word.istitle()
    last_word = last_word.lower()
    return rest_part, last_word, capitalized


e = inflect.engine()
re_camel_split = re.compile(r"^[a-z]+|[A-Z][^A-Z]*")


def plural_form(singular_form_name):
    rest_part, last_word, capitalized = split_last_word(singular_form_name)
    last_word = e.plural_noun(last_word)
    if capitalized:
        last_word = last_word.title()
    return rest_part + last_word


def singular_form(plural_form_name):
    rest_part, last_word, capitalized = split_last_word(plural_form_name)
    last_word = e.singular_noun(last_word)
    if capitalized:
        last_word = last_word.title()
    return rest_part + last_word


class TokenType(Enum):
    PrimitiveType = 1 << 0
    BeginList = 1 << 1
    VariableName = 1 << 2
    ClosingBracket = 1 << 3
    BeginDictionary = 1 << 4
    BeginStruct = 1 << 5
    BeginEnum = 1 << 6
    Number = 1 << 7
    Comma = 1 << 8


class Token:
    primitive_types = ['int', 'long', 'string', 'float', 'bool']
    value_to_token_types = {
        'List<': TokenType.BeginList,
        'Map<': TokenType.BeginDictionary,
        'Struct<': TokenType.BeginStruct,
        'Enum<': TokenType.BeginEnum,
        '>': TokenType.ClosingBracket,
        ',': TokenType.Comma,
    }
    re_variable_name = re.compile(r"^\w+\d*$")

    def __init__(self, value):
        self.value = value
        self.token_type = None
        if value in self.primitive_types:
            self.token_type = TokenType.PrimitiveType
            return
        self.token_type = self.value_to_token_types.get(value)
        if self.token_type is not None:
            return
        if str.isdigit(value):
            self.token_type = TokenType.Number
            return
        if self.re_variable_name.match(value):
            self.token_type = TokenType.VariableName
            return
        raise Exception(value)


class FieldParser:
    re_token = re.compile(r'[\d\w]+<?|[>,]')

    def __init__(self, table_name, field_name, field_def):
        self.table_name = table_name
        self.field_name = field_name
        self.field_def = field_def
        self.tokens = self.tokenize()
        self.cursor = 0
        self.field_info = self.parse_field_info(self.field_name)
        if self.cursor != len(self.tokens):
            raise Exception("")

    def tokenize(self):
        # print(self.field_def)
        return [Token(t) for t in self.re_token.findall(self.field_def)]

    def parse_field_info(self, field_name):
        if self.cursor >= len(self.tokens):
            raise Exception("")
        token = self.tokens[self.cursor]
        if token.token_type == TokenType.PrimitiveType:
            return self.parse_primitive(field_name)
        elif token.token_type == TokenType.BeginList:
            return self.parse_list(field_name)
        elif token.token_type == TokenType.BeginDictionary:
            return self.parse_dictionary(field_name)
        elif token.token_type == TokenType.BeginStruct:
            return self.parse_struct(field_name)
        elif token.token_type == TokenType.BeginEnum:
            return self.parse_enum(field_name)
        raise Exception(token.token_type)

    def parse_primitive(self, field_name):
        token = self.tokens[self.cursor]
        self.cursor += 1
        return FieldPrimitive(field_name, token.value)

    def parse_list(self, field_name):
        self.cursor += 1
        element_type = self.parse_field_info(singular_form(field_name))
        token = self.tokens[self.cursor]
        assert token.token_type == TokenType.ClosingBracket
        self.cursor += 1
        return FieldList(field_name, element_type)

    def parse_dictionary(self, field_name):
        self.cursor += 1
        key_type = self.parse_field_info('')
        token = self.tokens[self.cursor]
        assert token.token_type == TokenType.Comma
        self.cursor += 1
        value_type = self.parse_field_info('')
        token = self.tokens[self.cursor]
        assert token.token_type == TokenType.ClosingBracket
        self.cursor += 1
        return FieldDictionary(field_name, key_type, value_type)

    def parse_struct(self, field_name):
        self.cursor += 1
        struct_fields = []
        while True:
            field_info = self.parse_field_info('')
            struct_fields.append(field_info)
            token = self.tokens[self.cursor]
            assert token.token_type == TokenType.Comma
            self.cursor += 1
            token = self.tokens[self.cursor]
            assert token.token_type == TokenType.VariableName
            field_info.field_name = token.value
            self.cursor += 1
            token = self.tokens[self.cursor]
            if token.token_type == TokenType.ClosingBracket:
                break
            assert token.token_type == TokenType.Comma
            self.cursor += 1

        token = self.tokens[self.cursor]
        assert token.token_type == TokenType.ClosingBracket
        self.cursor += 1
        return FieldStruct(self.table_name, field_name, struct_fields)

    def parse_enum(self, field_name):
        self.cursor += 1
        enum_values = []
        while True:
            token = self.tokens[self.cursor]
            assert token.token_type == TokenType.VariableName
            enum_name = token.value
            self.cursor += 1
            token = self.tokens[self.cursor]
            assert token.token_type == TokenType.Comma
            self.cursor += 1
            token = self.tokens[self.cursor]
            assert token.token_type == TokenType.Number
            enum_value = token.value
            enum_values.append({"enum_name": enum_name, "enum_value": enum_value})
            self.cursor += 1
            token = self.tokens[self.cursor]
            if token.token_type == TokenType.ClosingBracket:
                break
            assert token.token_type == TokenType.Comma
            self.cursor += 1

        token = self.tokens[self.cursor]
        assert token.token_type == TokenType.ClosingBracket
        self.cursor += 1
        return FieldEnum(self.table_name, field_name, enum_values)

    def get_field_info(self):
        return self.field_info


class Field:
    def __init__(self, field_name, field_def):
        self.field_name = field_name
        self.field_def = field_def


class FieldPrimitive(Field):
    def __init__(self, field_name, field_type):
        super().__init__(field_name, field_type)


class FieldList(Field):
    def __init__(self, field_name, element_type):
        super().__init__(field_name, '')
        self.list_element_type = element_type


class FieldDictionary(Field):
    def __init__(self, field_name, key_type, value_type):
        super().__init__(field_name, '')
        self.dict_key_type = key_type
        self.dict_value_type = value_type


class FieldStruct(Field):
    def __init__(self, table_name, field_name, struct_fields):
        self.table_name = table_name
        field_def = upper_camel_case(field_name)
        if field_def == field_name:
            field_def = 'S' + field_name
        super().__init__(field_name, field_def)
        self.struct_fields = struct_fields


class FieldEnum(Field):
    def __init__(self, table_name, field_name, enum_values):
        self.table_name = table_name
        field_def = upper_camel_case(field_name)
        if field_def == field_name:
            field_def = 'E' + field_name
        super().__init__(field_name, field_def)
        self.enum_values = enum_values


class Scheme:

    def __init__(self, name, fields):
        self.full_name = name
        self.name = name
        self.fields = fields

    def get_associated_structs(self):
        structs = []
        for x in self.fields:
            if isinstance(x, FieldStruct):
                structs.append(x)
            elif isinstance(x, FieldList) and isinstance(x.list_element_type, FieldStruct):
                structs.append(x.list_element_type)
            elif isinstance(x, FieldDictionary):
                if isinstance(x.dict_value_type, FieldStruct):
                    structs.append(x.dict_value_type)
        return structs

    def get_associated_enums(self):
        enums = []
        for x in self.fields:
            if isinstance(x, FieldEnum):
                enums.append(x)
            elif isinstance(x, FieldList) and isinstance(x.list_element_type, FieldEnum):
                enums.append(x.list_element_type)
            elif isinstance(x, FieldDictionary):
                if isinstance(x.dict_key_type, FieldEnum):
                    enums.append(x.dict_key_type)
                if isinstance(x.dict_value_type, FieldEnum):
                    enums.append(x.dict_value_type)

        return enums


class Table:
    special_str = {
        '\n': ';l/~',
        '\\,': ':l/~',
    }

    def __init__(self, sheet):

        table_name = sheet.title
        self.name = table_name

        data = []
        self.scheme = None

        field_names, field_defs = None, None

        for row in sheet.iter_rows(values_only=True):

            if len(row) == 0:
                continue
            if row[0] is None:
                continue
            if self.scheme is None:
                if field_names is None:
                    field_names = self.try_read_field_names(row)
                    continue
                if field_defs is None:
                    field_defs = self.try_read_field_defs(row)
                    continue
                pairs = zip(field_names, field_defs)
                fields = [FieldParser(table_name, field_name, field_def).get_field_info()
                          for field_name, field_def in pairs]

                self.scheme = Scheme(table_name, fields)

            self.append_row(data, row)

        if self.scheme is None:
            raise Exception("invalid sheet file: " + sheet.title)
        if len(data) > 0:
            self.data = data

    def try_read_field_names(self, row):
        return (str(x) for x in row if x is not None and x != '')

    def try_read_field_defs(self, row):
        return (str(x) for x in row if x is not None and x != '')

    def to_safe_str(self, raw_str):
        if raw_str is None:
            return ''
        new_str = str(raw_str)
        for old, new in self.special_str.items():
            new_str = new_str.replace(old, new)
        return new_str

    def append_row(self, data, row):
        row_data = []
        count = len(self.scheme.fields)
        for index, cell in enumerate(row):
            if index < count:
                self.append_cell(row_data, cell, self.scheme.fields[index])
        data.append(row_data)

    def append_cell(self, row, cell, field):
        cell = self.to_safe_str(cell)
        if isinstance(field, FieldPrimitive):
            row.append(cell)
        elif isinstance(field, FieldList):
            if cell != '':
                divisions = str.count(cell, ',') + 1
                if isinstance(field.list_element_type, FieldStruct):
                    prefix = str(divisions // len(field.list_element_type.struct_fields))
                else:
                    prefix = str(divisions)
                row.append(prefix)
                row.append(cell)
            else:
                row.append('0')
        elif isinstance(field, FieldDictionary):
            if cell != '':
                divisions = str.count(cell, ',') + 1
                if isinstance(field.dict_value_type, FieldStruct):
                    prefix = str(divisions // (1 + len(field.dict_value_type.struct_fields)))
                else:
                    prefix = str(divisions // 2)
                row.append(prefix)
                row.append(cell)
            else:
                row.append('0')
        elif isinstance(field, FieldStruct):
            row.append(cell)
        elif isinstance(field, FieldEnum):
            row.append(cell)
        else:
            raise Exception('syntax error:' + field.field_def)


class TableReader:

    def __init__(self, *args, **kwargs):
        self.path_source = kwargs.get("source") or '.'

    def find_files(self):
        if self.path_source is None:
            self.path_source = os.getcwd()
        source_files = []
        for root, dirs, files in os.walk(self.path_source):
            source_files.extend([os.path.join(root, f) for f in files if f.endswith('.xlsx') and not f.startswith('_')])
        return source_files

    def create_tables(self):
        source_files = self.find_files()
        tables = []
        for file in source_files:
            wb = load_workbook(filename=file, read_only=True, data_only=True)
            for sheet in wb:
                if not sheet.title.startswith('_'):
                    print(f"reading {os.path.basename(file)}:[{sheet.title}]...")
                    tables.append(Table(sheet))
        tables.sort(key=lambda t: t.name)
        return tables


class TableWriter(ABC):

    def __init__(self, *args, **kwargs):
        self.path_out = kwargs.get("out") or './out'
        self.path_out_data = kwargs.get("outdata") or './out_data'
        self.name_space = kwargs.get("namespace") or 'EasyConverter'
        self.file_ext = self.get_script_file_ext()

    def ensure_path(self, path):
        file_dir = os.path.dirname(path)
        if not os.path.isdir(file_dir):
            os.makedirs(file_dir)

    def write_config(self, filename, text):
        path = f"{self.path_out}/{filename}"
        self.ensure_path(path)
        with open(path, "w", encoding='utf8') as f:
            f.write(text)

    def write_config_data(self, filename, text):
        path = f"{self.path_out_data}/{filename}"
        self.ensure_path(path)
        with open(path, "w", encoding='utf8') as f:
            f.write(text)

    def pack_table_data(self, table):
        data_arr = [str.join(",", row) for row in table.data]
        return str.join('\n', data_arr)

    @abstractmethod
    def get_script_file_ext(self):
        raise NotImplementedError()

    @abstractmethod
    def write_all(self, tables):
        raise NotImplementedError()


class IntegratedTableWriter(TableWriter):
    pass


class SeparatedTableWriter(TableWriter):
    pass


class EasyConverter:
    @staticmethod
    def convert(reader, writer):
        tables = reader.create_tables()
        writer.write_all(tables)

    @staticmethod
    def parse_args():
        parser = ArgumentParser()
        parser.add_argument("-source", type=str)
        parser.add_argument("-out", type=str, default='./out')
        parser.add_argument("-outdata", type=str, default='./out/data')
        parser.add_argument("-namespace", type=str, default='EasyConverter')
        return vars(parser.parse_args())
