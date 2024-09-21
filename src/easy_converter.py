#!/usr/bin/python
# -*- coding: UTF-8 -*-
import os
import re
from abc import ABC
from abc import abstractmethod
from datetime import datetime
from enum import Enum
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from argparse import ArgumentParser
import inflect
from jinja2 import Environment, FileSystemLoader
from typing import List, Tuple, Union, Any, Optional, Iterator
from functools import partial


def upper_camel_case(var_name: str) -> str:
    if var_name == '':
        return var_name
    return var_name[0].upper() + var_name[1:]


def split_last_word(var_name: str) -> Tuple[str, Any, Any]:
    splits = re_camel_split.findall(var_name)
    last_word = splits[-1]
    rest_part = ""
    if len(splits) > 1:
        rest_part = "".join(splits[0:-1])
    capitalized = last_word.istitle()
    last_word = last_word.lower()
    return rest_part, last_word, capitalized


engine = inflect.engine()
re_camel_split = re.compile(r"^[a-z]+|[A-Z][^A-Z]*")


def plural_form(singular_form_name):
    rest_part, last_word, capitalized = split_last_word(singular_form_name)
    last_word = engine.plural_noun(last_word)
    if capitalized:
        last_word = last_word.title()
    return rest_part + last_word


def singular_form(plural_form_name):
    rest_part, last_word, capitalized = split_last_word(plural_form_name)
    singular = engine.singular_noun(last_word)
    if singular is not False:
        last_word = singular
    if capitalized:
        last_word = last_word.title()
    return rest_part + last_word


class TokenType(Enum):
    PrimitiveType = 1 << 0
    BeginList = 1 << 1
    VariableName = 1 << 2
    ClosingBracket = 1 << 3
    BeginDictionary = 1 << 4
    BeginStruct = 1 << 5
    BeginEnum = 1 << 6
    Number = 1 << 7
    Comma = 1 << 8
    Reference = 1 << 9


class Token:
    primitive_types = ['int', 'long', 'string', 'float', 'bool']
    value_to_token_types = {
        'List<': TokenType.BeginList,
        'Map<': TokenType.BeginDictionary,
        'Struct<': TokenType.BeginStruct,
        'Enum<': TokenType.BeginEnum,
        '>': TokenType.ClosingBracket,
        ',': TokenType.Comma,
        '@': TokenType.Reference,
    }
    re_variable_name = re.compile(r"^\w+\d*$")

    def __init__(self, value: str):
        self.value: str = value
        self.token_type: Optional[TokenType] = None
        if value in self.primitive_types:
            self.token_type = TokenType.PrimitiveType
            return
        self.token_type = self.value_to_token_types.get(value)
        if self.token_type is not None:
            return
        if str.isdigit(value):
            self.token_type = TokenType.Number
            return
        if self.re_variable_name.match(value):
            self.token_type = TokenType.VariableName
            return
        raise Exception(value)


class EnumFieldValue:
    def __init__(self, enum_name, enum_value):
        self.enum_name = enum_name
        self.enum_value = enum_value


class FieldParser:
    re_token = re.compile(r'\w+<?|[>,@]')

    def __init__(self, table_name: str, field_name: str, field_def: str):
        self.table_name: str = table_name
        self.field_name: str = field_name
        self.field_def: str = field_def
        self.tokens: List[Token] = self.tokenize()
        self.cursor: int = 0
        self.__field_info: Field = self.parse_field_info(self.field_name)
        if self.cursor != len(self.tokens):
            raise Exception("")

    def tokenize(self) -> List[Token]:
        # print(self.field_def)
        return [Token(t) for t in self.re_token.findall(self.field_def)]

    def parse_field_info(self, field_name: str) -> 'Field':
        if self.cursor >= len(self.tokens):
            raise Exception("")
        token = self.tokens[self.cursor]
        if token.token_type == TokenType.PrimitiveType:
            return self.parse_primitive(field_name)
        elif token.token_type == TokenType.BeginList:
            return self.parse_list(field_name)
        elif token.token_type == TokenType.BeginDictionary:
            return self.parse_dictionary(field_name)
        elif token.token_type == TokenType.BeginStruct:
            return self.parse_struct(field_name)
        elif token.token_type == TokenType.BeginEnum:
            return self.parse_enum(field_name)
        elif token.token_type == TokenType.Reference:
            return self.parse_reference(field_name)
        raise Exception(token.token_type)

    def parse_primitive(self, field_name: str) -> 'FieldPrimitive':
        token = self.tokens[self.cursor]
        self.cursor += 1
        return FieldPrimitive(self.table_name, field_name, token.value)

    def parse_list(self, field_name: str) -> 'FieldList':
        self.cursor += 1
        element_type = self.parse_field_info(singular_form(field_name))
        token = self.tokens[self.cursor]
        assert token.token_type == TokenType.ClosingBracket
        self.cursor += 1
        return FieldList(self.table_name, field_name, element_type)

    def parse_dictionary(self, field_name: str) -> 'FieldDictionary':
        self.cursor += 1
        key_type = self.parse_field_info('')
        token = self.tokens[self.cursor]
        assert token.token_type == TokenType.Comma
        self.cursor += 1
        value_type = self.parse_field_info('')
        token = self.tokens[self.cursor]
        assert token.token_type == TokenType.ClosingBracket
        self.cursor += 1
        return FieldDictionary(self.table_name, field_name, key_type, value_type)

    def parse_struct(self, field_name: str) -> 'FieldStruct':
        self.cursor += 1
        struct_fields = []
        while True:
            field_info = self.parse_field_info('')
            struct_fields.append(field_info)
            token = self.tokens[self.cursor]
            assert token.token_type == TokenType.Comma
            self.cursor += 1
            token = self.tokens[self.cursor]
            assert token.token_type == TokenType.VariableName
            field_info.field_name = token.value
            self.cursor += 1
            token = self.tokens[self.cursor]
            if token.token_type == TokenType.ClosingBracket:
                break
            assert token.token_type == TokenType.Comma
            self.cursor += 1

        token = self.tokens[self.cursor]
        assert token.token_type == TokenType.ClosingBracket
        self.cursor += 1
        return FieldStruct(self.table_name, field_name, struct_fields)

    def parse_enum(self, field_name: str) -> 'FieldEnum':
        self.cursor += 1
        enum_values = []
        while True:
            token = self.tokens[self.cursor]
            assert token.token_type == TokenType.VariableName
            enum_name = token.value
            self.cursor += 1
            token = self.tokens[self.cursor]
            assert token.token_type == TokenType.Comma
            self.cursor += 1
            token = self.tokens[self.cursor]
            assert token.token_type == TokenType.Number
            enum_value = token.value
            enum_values.append(EnumFieldValue(enum_name, enum_value))
            self.cursor += 1
            token = self.tokens[self.cursor]
            if token.token_type == TokenType.ClosingBracket:
                break
            assert token.token_type == TokenType.Comma
            self.cursor += 1

        token = self.tokens[self.cursor]
        assert token.token_type == TokenType.ClosingBracket
        self.cursor += 1
        return FieldEnum(self.table_name, field_name, enum_values)

    def parse_reference(self, field_name: str) -> 'FieldReference':
        self.cursor += 1
        token = self.tokens[self.cursor]
        assert token.token_type == TokenType.VariableName
        ref_table_name = token.value
        self.cursor += 1
        token = self.tokens[self.cursor]
        assert token.token_type == TokenType.Comma
        self.cursor += 1
        token = self.tokens[self.cursor]
        assert token.token_type == TokenType.VariableName
        ref_type_name = token.value
        self.cursor += 1
        return FieldReference(self.table_name, field_name, ref_table_name, ref_type_name)

    @property
    def field_info(self):
        return self.__field_info


class FieldType(Enum):
    Primitive = 1
    List = 2
    Dictionary = 3
    Struct = 4
    Enum = 5
    Reference = 6


class Field:
    def __init__(self, table_name: str, field_name: str, field_def: str):
        self.table_name: str = table_name
        self.field_name: str = field_name
        self.field_def: str = field_def
        self.field_type: Optional[FieldType] = None

    def get_associated_structs(self) -> Iterator["FieldStruct"]:
        yield from ()

    def get_associated_enums(self) -> Iterator["FieldEnum"]:
        yield from ()

    def get_associated_references(self) -> Iterator["FieldReference"]:
        yield from ()

    def cell_to_str(self, raw_cell: Union[str, float, datetime, None]) -> str:
        return CellData.to_safe_str(raw_cell)


class FieldPrimitive(Field):
    def __init__(self, table_name: str, field_name: str, field_type: str):
        super().__init__(table_name, field_name, field_type)
        self.field_type = FieldType.Primitive


class FieldList(Field):
    def __init__(self, table_name: str, field_name: str, element_type: Field):
        super().__init__(table_name, field_name, '')
        self.field_type = FieldType.List
        self.list_element_type = element_type

    def get_associated_structs(self) -> Iterator["FieldStruct"]:
        yield from self.list_element_type.get_associated_structs()

    def get_associated_enums(self) -> Iterator["FieldEnum"]:
        yield from self.list_element_type.get_associated_enums()

    def get_associated_references(self) -> Iterator["FieldReference"]:
        yield from self.list_element_type.get_associated_references()

    def cell_to_str(self, raw_cell: Union[str, float, datetime, None]) -> str:
        if raw_cell != '':
            elements = CellData.cell_to_elements(raw_cell, self)
            raw_str = str(len(elements))
            for ele in elements:
                raw_str = raw_str + ','
                raw_str = raw_str + self.list_element_type.cell_to_str(ele)
            return raw_str
        else:
            return '0'


class FieldDictionary(Field):
    def __init__(self, table_name: str, field_name: str, key_type: Field, value_type: Field):
        super().__init__(table_name, field_name, '')
        self.field_type = FieldType.Dictionary
        self.dict_key_type: Field = key_type
        self.dict_value_type: Field = value_type

    def get_associated_structs(self) -> Iterator["FieldStruct"]:
        yield from self.dict_key_type.get_associated_structs()
        yield from self.dict_value_type.get_associated_structs()

    def get_associated_enums(self) -> Iterator["FieldEnum"]:
        yield from self.dict_key_type.get_associated_enums()
        yield from self.dict_value_type.get_associated_enums()

    def get_associated_references(self) -> Iterator["FieldReference"]:
        yield from self.dict_key_type.get_associated_references()
        yield from self.dict_value_type.get_associated_references()

    def cell_to_str(self, raw_cell: Union[str, float, datetime, None]) -> str:
        if raw_cell != '':
            elements = CellData.cell_to_elements(raw_cell, self)
            raw_str = str(len(elements))
            for (k, v) in elements:
                raw_str = raw_str + ','
                raw_str = raw_str + self.dict_key_type.cell_to_str(k)
                raw_str = raw_str + ','
                raw_str = raw_str + self.dict_value_type.cell_to_str(v)
            return raw_str
        else:
            return '0'


class FieldStruct(Field):
    def __init__(self, table_name: str, field_name: str, struct_fields: List[Field]):
        field_def = upper_camel_case(field_name)
        if field_def == field_name:
            field_def = 'S' + field_name
        super().__init__(table_name, field_name, field_def)
        self.field_type = FieldType.Struct
        self.struct_fields: List[Field] = struct_fields

    def get_associated_structs(self) -> Iterator["FieldStruct"]:
        yield self
        for f in self.struct_fields:
            yield from f.get_associated_structs()

    def get_associated_enums(self) -> Iterator["FieldEnum"]:
        for f in self.struct_fields:
            yield from f.get_associated_enums()

    def get_associated_references(self) -> Iterator["FieldReference"]:
        for f in self.struct_fields:
            yield from f.get_associated_references()

    def cell_to_str(self, raw_cell: Union[str, float, datetime, None]) -> str:
        raw_str = ''
        elements = CellData.cell_to_elements(raw_cell, self)
        for f, e in zip(self.struct_fields, elements):
            if raw_str != '':
                raw_str = raw_str + ','
            raw_str = raw_str + f.cell_to_str(e)
        return raw_str


class FieldEnum(Field):
    def __init__(self, table_name: str, field_name: str, enum_values: List[EnumFieldValue]):
        self.table_name = table_name
        field_def = upper_camel_case(field_name)
        if field_def == field_name:
            field_def = 'E' + field_name
        super().__init__(table_name, field_name, field_def)
        self.field_type = FieldType.Enum
        self.enum_values = enum_values
        self.enum_key_to_value = {item.enum_name: item.enum_value for item in enum_values}

    def get_associated_enums(self) -> Iterator["FieldEnum"]:
        yield self

    def cell_to_str(self, raw_cell: Union[str, float, datetime, None]) -> str:
        enum_value = self.enum_key_to_value.get(raw_cell)
        if enum_value is None:
            raise Exception('invalid enum key:' + str(raw_cell))
        else:
            return str(enum_value)


class FieldReference(Field):
    def __init__(self, table_name: str, field_name: str, ref_table_name: str, ref_type_name: str):
        super().__init__(table_name, field_name, '')
        self.field_type = FieldType.Reference
        self.ref_table_name: str = ref_table_name
        self.ref_type_name: str = ref_type_name
        self.ref_type: Optional[Field] = None

    def get_associated_references(self) -> Iterator["FieldReference"]:
        yield self

    def cell_to_str(self, raw_cell: Union[str, float, datetime, None]) -> str:
        assert self.ref_type is not None, \
            f'Reference not found: {self.ref_table_name},{self.ref_type_name} in {self.table_name},{self.field_name}'
        return self.ref_type.cell_to_str(raw_cell)


class Scheme:

    def __init__(self, name: str, field_pairs: Iterator[Tuple[str, str]]):
        self.full_name: str = name
        self.name: str = name
        self.fields: List[Field] = self.populate_fields(field_pairs)

    def populate_fields(self, field_pairs: Iterator[Tuple[str, str]]) -> List[Field]:
        fields: List[Field] = []
        for field_name, field_def in field_pairs:
            parser = FieldParser(self.name, field_name, field_def)
            fields.append(parser.field_info)
        return fields

    def get_associated_structs(self) -> List["FieldStruct"]:
        structs = []
        for f in self.fields:
            structs.extend(f.get_associated_structs())
        return structs

    def get_associated_enums(self) -> List['FieldEnum']:
        enums = []
        for f in self.fields:
            enums.extend(f.get_associated_enums())
        return enums

    def get_associated_references(self) -> List['FieldReference']:
        references = []
        for f in self.fields:
            references.extend(f.get_associated_references())
        return references


class CellData:
    def __init__(self, raw_cell: Union[str, float, datetime, None], field: Field):
        self.field = field
        self.raw_str = field.cell_to_str(raw_cell)

    @staticmethod
    def cell_to_elements(raw_cell, field):
        try:
            if raw_cell is None or raw_cell == '':
                return []
            if isinstance(field, FieldList):
                elements = raw_cell.split(',')
                sub_element_size = 1
                if isinstance(field.list_element_type, FieldStruct):
                    sub_element_size = len(field.list_element_type.struct_fields)
                return [','.join(elements[i:i + sub_element_size])
                        for i in range(0, len(elements), sub_element_size)]
            elif isinstance(field, FieldDictionary):
                elements = raw_cell.split(',')
                sub_element_size = 1
                if isinstance(field.dict_value_type, FieldStruct):
                    sub_element_size = len(field.dict_value_type.struct_fields)
                return [(elements[i], ','.join(elements[i + 1: i + 1 + sub_element_size]))
                        for i in range(0, len(elements), sub_element_size + 1)]
            elif isinstance(field, FieldStruct):
                elements = raw_cell.split(',')
                return elements
            return [raw_cell]
        except Exception as e:
            print(f"error in reading: {field.table_name},{field.field_name},{raw_cell}")
            raise e

    @staticmethod
    def to_safe_str(raw_str: Union[str, float, datetime, None]) -> str:
        if raw_str is None:
            return ''
        new_str = str(raw_str)
        for old, new in Table.special_str.items():
            new_str = new_str.replace(old, new)
        return new_str

    def __str__(self) -> str:
        if self.raw_str is None:
            return ''
        return self.raw_str


class RowData:
    def __init__(self):
        self.cells: List[CellData] = []

    def append(self, cell: CellData):
        self.cells.append(cell)

    def __iter__(self) -> Iterator[CellData]:
        return (cell for cell in self.cells)


class Table:
    special_str = {
        '\n': ';l/~',
        '\\,': ':l/~',
    }

    def __init__(self, sheet: Worksheet):
        self.__sheet: Worksheet = sheet
        self.name = upper_camel_case(sheet.title)

        field_names = self.try_read_field_names()
        field_defs = self.try_read_field_defs()

        pairs: Iterator[Tuple[str, str]] = zip(field_names, field_defs)
        self.scheme: Scheme = Scheme(self.name, pairs)

        self.data: List[RowData] = []

    def try_read_field_names(self) -> Iterator[str]:
        for row in self.__sheet.iter_rows(values_only=True, min_row=1, max_row=1):
            return (str(x) for x in row if x is not None and x != '')

    def try_read_field_defs(self) -> Iterator[str]:
        for row in self.__sheet.iter_rows(values_only=True, min_row=2, max_row=2):
            return (str(x) for x in row if x is not None and x != '')

    def populate_table_data(self) -> None:
        columns = len(self.scheme.fields)
        for row in self.__sheet.iter_rows(values_only=True, min_row=3, min_col=1, max_col=columns):
            if row[0] is None:
                continue
            row_data: RowData = RowData()
            for index, cell in enumerate(row):
                cell_data = CellData(cell, self.scheme.fields[index])
                row_data.append(cell_data)
            self.data.append(row_data)


class TableReader:

    def __init__(self, *args, **kwargs):
        self.path_source = kwargs.get("source") or '.'

    def find_files(self):
        if self.path_source is None:
            self.path_source = os.getcwd()
        source_files = []
        for root, dirs, files in os.walk(self.path_source):
            source_files.extend([os.path.join(root, f) for f in files if f.endswith('.xlsx') and not f.startswith('_')])
        return source_files

    def create_tables(self) -> List[Table]:
        source_files = self.find_files()
        tables = []
        for file in source_files:
            wb = load_workbook(filename=file, read_only=True, data_only=True)
            for sheet in wb:
                if not sheet.title.startswith('_'):
                    print(f"reading {os.path.basename(file)}:[{sheet.title}]...")
                    tables.append(Table(sheet))
        tables.sort(key=lambda t: t.name)
        self.process_reference_types(tables)
        for table in tables:
            table.populate_table_data()
        return tables

    def process_reference_types(self, tables: List[Table]):
        sub_types = {}
        for table in tables:
            structs = table.scheme.get_associated_structs()
            for s in structs:
                full_name = s.table_name + "," + s.field_def
                sub_types[full_name] = s
            enums = table.scheme.get_associated_enums()
            for e in enums:
                full_name = e.table_name + "," + e.field_def
                sub_types[full_name] = e
        for table in tables:
            references = table.scheme.get_associated_references()
            for reference in references:
                ref_type_full_name = reference.ref_table_name + "," + reference.ref_type_name
                ref_type = sub_types.get(ref_type_full_name)
                reference.ref_type = ref_type


class TableWriter(ABC):

    def __init__(self, *args, **kwargs):
        self.path_out: str = kwargs.get("out") or './out'
        self.path_out_data: str = kwargs.get("outdata") or './out_data'
        self.name_space: str = kwargs.get("namespace") or 'EasyConverter'
        self.file_ext: str = self.get_script_file_ext()
        python_file_path = os.path.abspath(__file__)
        python_dir_path = os.path.dirname(python_file_path)
        template_path = os.path.join(python_dir_path, self.get_template_file_dir())
        self.env: Environment = Environment(
            loader=FileSystemLoader(template_path),
            trim_blocks=True,
            lstrip_blocks=True
        )
        self.env.globals['FieldType'] = FieldType
        self.env.globals['name_space'] = self.name_space
        self.env.filters['upper_camel_case'] = upper_camel_case
        self.env.filters['plural_form'] = plural_form
        self.env.globals['get_display_def'] = partial(self.get_display_def)

    @staticmethod
    def ensure_path(path: str):
        file_dir = os.path.dirname(path)
        if not os.path.isdir(file_dir):
            os.makedirs(file_dir)

    def write_config(self, filename: str, text: str):
        path = f"{self.path_out}/{filename}"
        self.ensure_path(path)
        with open(path, "w", encoding='utf8') as f:
            f.write(text)

    def write_config_data(self, filename: str, text: str):
        path = f"{self.path_out_data}/{filename}"
        self.ensure_path(path)
        with open(path, "w", encoding='utf8') as f:
            f.write(text)

    @staticmethod
    def pack_table_data(table: Table) -> str:
        rows = (str.join(",", (str(cell) for cell in row)) for row in table.data)
        return str.join('\n', rows)

    @abstractmethod
    def get_template_file_dir(self) -> str:
        raise NotImplementedError()

    @abstractmethod
    def get_script_file_ext(self) -> str:
        raise NotImplementedError()

    @abstractmethod
    def get_display_def(self, filed: Field) -> str:
        raise NotImplementedError()

    @abstractmethod
    def write_all(self, tables: List[Table]):
        raise NotImplementedError()


class EasyConverter:
    @staticmethod
    def convert(reader: TableReader, writer: TableWriter):
        tables = reader.create_tables()
        writer.write_all(tables)

    @staticmethod
    def parse_args():
        parser = ArgumentParser()
        parser.add_argument("-source", type=str)
        parser.add_argument("-out", type=str, default='./out')
        parser.add_argument("-outdata", type=str, default='./out/data')
        parser.add_argument("-namespace", type=str, default='EasyConverter')
        return vars(parser.parse_args())
